#!/usr/bin/env python3
"""Create Chinese-friendly Markdown copies and formatted XLSX review sheets."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[3]
HUMAN_ROOT = ROOT / "aaai27/human_evaluation"

TASKS = {
    "physical": ("物理属性预测", "physical_property_expert"),
    "indoor3d": ("室内单图 3D 场景生成", "indoor3d_expert"),
    "iad": ("工业异常检测 IAD + Agent", "iad_expert"),
}

DIMS = [
    ("创新性", "novelty"), ("研究吸引力", "excitement"), ("可行性", "feasibility"),
    ("预期有效性", "expected_effectiveness"), ("综合质量", "overall"),
    ("基线依据充分性", "baseline_grounding"), ("实验严谨性", "experimental_rigor"),
    ("机制具体性", "mechanism_specificity"), ("实现就绪度", "implementation_readiness"),
]

BLUE = "D9EAF7"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
GRAY = "E7E6E6"
RED = "FCE4D6"
THIN = Side(style="thin", color="B7B7B7")


def read_jsonl(path: Path) -> list[dict]:
    return [json.loads(x) for x in path.open(encoding="utf-8") if x.strip()]


def chinese_markdown(source: str) -> str:
    return (
        source.replace("## Item ", "## 条目 ")
        .replace("类型：`single_idea`", "类型：`单个 Idea 对比`")
        .replace("类型：`portfolio`", "类型：`三 Idea 组合对比`")
        .replace("### Candidate A", "### 候选方案 A")
        .replace("### Candidate B", "### 候选方案 B")
        .replace("ANONYMOUS_REVIEW_PACKET.md", "匿名评审材料_中文版.md")
        .replace("RESPONSE_SHEET.csv", "评审答题表_中文版.xlsx")
    )


def style_header(ws, row: int, start: int, end: int, fill: str) -> None:
    for cell in ws.iter_cols(min_col=start, max_col=end, min_row=row, max_row=row):
        c = cell[0]
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(task: str, label: str, reviewer_code: str, items: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "评分表"
    headers = ["评审者代码", "序号", "条目ID", "对比类型", "子问题熟悉度（1-5）"]
    for cn, _ in DIMS:
        headers.extend([f"{cn}-A（1-5）", f"{cn}-B（1-5）"])
    headers.extend(["总体偏好（A/B/tie）", "置信度（1-5）", "选择理由（必填，1-3句）", "疑虑或共同缺陷（选填）", "用时（分钟）"])
    ws.append(headers)
    for idx, item in enumerate(items, 1):
        kind = "三 Idea 组合" if item["family_public"] == "portfolio" else "单个 Idea"
        ws.append([reviewer_code, idx, item["item_id"], kind] + [""] * (len(headers) - 4))

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).coordinate}"
    ws.row_dimensions[1].height = 46
    widths = {"A": 27, "B": 7, "C": 18, "D": 14, "E": 19}
    for col, width in widths.items(): ws.column_dimensions[col].width = width
    for col in range(6, 24): ws.column_dimensions[ws.cell(1, col).column_letter].width = 16
    ws.column_dimensions[ws.cell(1, 24).column_letter].width = 20
    ws.column_dimensions[ws.cell(1, 25).column_letter].width = 16
    ws.column_dimensions[ws.cell(1, 26).column_letter].width = 46
    ws.column_dimensions[ws.cell(1, 27).column_letter].width = 42
    ws.column_dimensions[ws.cell(1, 28).column_letter].width = 13
    style_header(ws, 1, 1, 5, GRAY)
    style_header(ws, 1, 6, 23, BLUE)
    style_header(ws, 1, 24, 28, GREEN)
    for row in ws.iter_rows(min_row=2, max_row=21, min_col=1, max_col=28):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.row_dimensions[row[0].row].height = 32

    score_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    score_validation.error = "请输入 1–5 的整数。"; score_validation.errorTitle = "评分格式错误"; score_validation.showErrorMessage = True
    pref_validation = DataValidation(type="list", formula1='"A,B,tie"', allow_blank=True)
    pref_validation.error = "请选择 A、B 或 tie。"; pref_validation.showErrorMessage = True
    ws.add_data_validation(score_validation); ws.add_data_validation(pref_validation)
    score_validation.add("E2:W21"); score_validation.add("Y2:Y21")
    pref_validation.add("X2:X21")
    ws.conditional_formatting.add("Z2:Z21", FormulaRule(formula=["LEN(Z2)=0"], fill=PatternFill("solid", fgColor=RED)))

    content = wb.create_sheet("候选内容")
    content.append(["序号", "条目ID", "对比类型", "候选方案 A（科研原文）", "候选方案 B（科研原文）"])
    for idx, item in enumerate(items, 1):
        kind = "三 Idea 组合" if item["family_public"] == "portfolio" else "单个 Idea"
        if len(item["candidate_a"]) > 32767 or len(item["candidate_b"]) > 32767:
            raise ValueError(f"{item['item_id']} exceeds Excel cell text limit; use Markdown packet")
        content.append([idx, item["item_id"], kind, item["candidate_a"], item["candidate_b"]])
    content.freeze_panes = "D2"
    content.auto_filter.ref = "A1:E21"
    for col, width in zip("ABCDE", (7, 18, 14, 75, 75)): content.column_dimensions[col].width = width
    style_header(content, 1, 1, 3, GRAY); style_header(content, 1, 4, 5, YELLOW)
    for row in content.iter_rows(min_row=2, max_row=21, min_col=1, max_col=5):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        content.row_dimensions[row[0].row].height = 220

    guide = wb.create_sheet("评分说明")
    guide.append([f"{label}匿名科研 Idea 盲评", "说明"])
    guide.append(["评审者代码", reviewer_code])
    guide.append(["核心纪律", "独立完成；不使用大模型代评；不猜测方法来源；不因篇幅更长直接判胜；tie 是合法选项。"])
    guide.append(["填写顺序", "先在“候选内容”阅读同一条目的 A/B，再回到“评分表”填写相同条目ID的一行。"])
    guide.append(["分数", "1=明显不足；2=较弱；3=基本合理；4=较强；5=非常强。所有评分必须为整数。"])
    for cn, en in DIMS:
        explanations = {
            "novelty": "核心机制是否真正区别于 baseline 拼接，而非换名或堆工具。",
            "excitement": "若结果成立，是否值得研究社区关注。",
            "feasibility": "数据、标签、算力与时间是否现实。",
            "expected_effectiveness": "机制是否有合理路径改善目标指标。",
            "overall": "综合科学价值，而非写作长度或语言流畅度。",
            "baseline_grounding": "是否指出具体 baseline 边界，机制是否直接针对该缺陷。",
            "experimental_rigor": "baseline、消融、negative control、指标和可证伪条件是否充分。",
            "mechanism_specificity": "输入、输出、算法步骤、目标函数或决策规则是否明确一致。",
            "implementation_readiness": "能否据此开始实现；数据、脚本和产物是否合理。",
        }
        guide.append([cn, explanations[en]])
    guide.append(["总体偏好", "只能选择 A、B 或 tie；综合判断，不要机械求和。"])
    guide.append(["理由", "每条至少 1–3 句，指出机制、实验、数据或可实现性的具体依据。"])
    guide.column_dimensions["A"].width = 24; guide.column_dimensions["B"].width = 110
    style_header(guide, 1, 1, 2, GREEN)
    for row in guide.iter_rows(min_row=2, max_row=guide.max_row, min_col=1, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        row[0].font = Font(bold=True)
        guide.row_dimensions[row[0].row].height = 38
    guide.freeze_panes = "A2"

    wb.properties.title = f"{label}匿名科研 Idea 人类评审表"
    wb.properties.subject = "AAAI-27 human blind review"
    wb.save(path)


def main() -> None:
    for task, (label, reviewer_code) in TASKS.items():
        task_dir = HUMAN_ROOT / task
        items = read_jsonl(task_dir / "public_human_items.jsonl")
        source = (task_dir / "ANONYMOUS_REVIEW_PACKET.md").read_text(encoding="utf-8")
        (task_dir / "匿名评审材料_中文版.md").write_text(chinese_markdown(source), encoding="utf-8")
        build_workbook(task, label, reviewer_code, items, task_dir / "评审答题表_中文版.xlsx")
        print(f"Wrote Chinese review files for {task}")


if __name__ == "__main__":
    main()
