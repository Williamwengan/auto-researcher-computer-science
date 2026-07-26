#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


HEADERS = {
    "reviewer": ["评审者代码"],
    "item_id": ["条目ID"],
    "comparison_public": ["对比类型"],
    "familiarity": ["子问题熟悉度（1-5）", "子问题熟悉度（1-3）"],
    "preference": ["总体偏好（A/B/tie）"],
    "confidence": ["置信度（1-5）", "置信度（1-3）"],
    "rationale": ["选择理由（必填，1-3句）"],
    "concerns": ["疑虑或共同缺陷（选填）"],
}

DIMENSIONS = {
    "novelty": (["创新性-A（1-5）", "创新性-A（1-3）"], ["创新性-B（1-5）", "创新性-B（1-3）"]),
    "excitement": (["研究吸引力-A（1-5）", "研究吸引力-A（1-3）"], ["研究吸引力-B（1-5）", "研究吸引力-B（1-3）"]),
    "feasibility": (["可行性-A（1-5）", "可行性-A（1-3）"], ["可行性-B（1-5）", "可行性-B（1-3）"]),
    "expected_effectiveness": (["预期有效性-A（1-5）", "预期有效性-A（1-3）"], ["预期有效性-B（1-5）", "预期有效性-B（1-3）"]),
    "overall": (["综合质量-A（1-5）", "综合质量-A（1-3）"], ["综合质量-B（1-5）", "综合质量-B（1-3）"]),
    "baseline_grounding": (["基线依据充分性-A（1-5）", "基线依据充分性-A（1-3）"], ["基线依据充分性-B（1-5）", "基线依据充分性-B（1-3）"]),
    "experimental_rigor": (["实验严谨性-A（1-5）", "实验严谨性-A（1-3）"], ["实验严谨性-B（1-5）", "实验严谨性-B（1-3）"]),
    "mechanism_specificity": (["机制具体性-A（1-5）", "机制具体性-A（1-3）"], ["机制具体性-B（1-5）", "机制具体性-B（1-3）"]),
    "implementation_readiness": (["实现就绪度-A（1-5）", "实现就绪度-A（1-3）"], ["实现就绪度-B（1-5）", "实现就绪度-B（1-3）"]),
}


def read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def normalize_preference(value) -> str:
    text = str(value or "").strip().lower()
    if text in {"a", "b", "tie"}:
        return text
    if text in {"平局", "都可以", "相当"}:
        return "tie"
    return text


def as_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        return None


def resolve_header(header_to_idx: dict, candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in header_to_idx:
            return candidate
    return None


def wilson_ci(wins: float, n: int, z: float = 1.96) -> tuple[float, float]:
    if n == 0:
        return (0.0, 0.0)
    p = wins / n
    denom = 1 + z * z / n
    center = (p + z * z / (2 * n)) / denom
    margin = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n) / denom
    return (max(0.0, center - margin), min(1.0, center + margin))


def read_workbook(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    if "评分表" not in wb.sheetnames:
        raise ValueError(f"{path} missing sheet: 评分表")
    ws = wb["评分表"]
    raw_headers = [cell.value for cell in ws[1]]
    header_to_idx = {header: idx for idx, header in enumerate(raw_headers) if header}

    resolved_headers = {}
    missing = []
    for key, candidates in HEADERS.items():
        header = resolve_header(header_to_idx, candidates)
        if header is None:
            missing.append("/".join(candidates))
        else:
            resolved_headers[key] = header
    resolved_dims = {}
    for dim, (a_candidates, b_candidates) in DIMENSIONS.items():
        a_header = resolve_header(header_to_idx, a_candidates)
        b_header = resolve_header(header_to_idx, b_candidates)
        if a_header is None:
            missing.append("/".join(a_candidates))
        if b_header is None:
            missing.append("/".join(b_candidates))
        if a_header and b_header:
            resolved_dims[dim] = (a_header, b_header)
    if missing:
        raise ValueError(f"{path} missing columns: {missing}")

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in raw):
            continue
        row = {key: raw[header_to_idx[header]] for key, header in resolved_headers.items()}
        row["preference"] = normalize_preference(row["preference"])
        row["source_workbook"] = str(path)
        row["review_protocol_note"] = ""
        if "AI辅助说明" in wb.sheetnames:
            row["review_protocol_note"] = "ai_assisted_informal_review"
        row["dimension_scores"] = {}
        for dim, (a_header, b_header) in resolved_dims.items():
            row["dimension_scores"][dim] = {
                "A": as_float(raw[header_to_idx[a_header]]),
                "B": as_float(raw[header_to_idx[b_header]]),
            }
        rows.append(row)
    return rows


def decode_rows(rows: list[dict], answer_key: dict[str, dict]) -> tuple[list[dict], list[str]]:
    decoded = []
    errors = []
    seen = set()
    for row in rows:
        item_id = row.get("item_id")
        if not item_id:
            errors.append(f"missing item_id in {row['source_workbook']}")
            continue
        if item_id in seen:
            errors.append(f"duplicate item_id: {item_id}")
        seen.add(item_id)
        key = answer_key.get(item_id)
        if not key:
            errors.append(f"item_id absent from private key: {item_id}")
            continue
        pref = row["preference"]
        if pref not in {"a", "b", "tie"}:
            errors.append(f"{item_id} invalid preference: {pref}")
            continue

        mapping = key["candidate_mapping"]
        if pref == "tie":
            preferred_method = "tie"
            focal_result = "tie"
        else:
            side = pref.upper()
            preferred_method = mapping.get(side)
            focal_result = "win" if preferred_method == "focused_full" else "loss"

        dimension_deltas = {}
        focal_side = None
        for side, method in mapping.items():
            if method == "focused_full":
                focal_side = side
        for dim, scores in row["dimension_scores"].items():
            if focal_side in {"A", "B"}:
                other_side = "B" if focal_side == "A" else "A"
                if scores.get(focal_side) is not None and scores.get(other_side) is not None:
                    dimension_deltas[dim] = scores[focal_side] - scores[other_side]
                else:
                    dimension_deltas[dim] = None

        decoded.append(
            {
                **row,
                "task": key.get("task"),
                "family": key.get("family"),
                "private_comparison": key.get("private_comparison"),
                "candidate_mapping": mapping,
                "preferred_method": preferred_method,
                "focused_full_side": focal_side,
                "focused_full_result": focal_result,
                "dimension_deltas_focused_full_minus_other": dimension_deltas,
            }
        )
    return decoded, errors


def summarize_group(rows: list[dict]) -> dict:
    n = len(rows)
    wins = sum(1 for row in rows if row["focused_full_result"] == "win")
    losses = sum(1 for row in rows if row["focused_full_result"] == "loss")
    ties = sum(1 for row in rows if row["focused_full_result"] == "tie")
    tie_half_wins = wins + 0.5 * ties
    low, high = wilson_ci(tie_half_wins, n)
    confidence_values = [as_float(row.get("confidence")) for row in rows]
    confidence_values = [v for v in confidence_values if v is not None]
    familiarity_values = [as_float(row.get("familiarity")) for row in rows]
    familiarity_values = [v for v in familiarity_values if v is not None]
    dim_means = {}
    for dim in DIMENSIONS:
        vals = [row["dimension_deltas_focused_full_minus_other"].get(dim) for row in rows]
        vals = [v for v in vals if v is not None]
        dim_means[dim] = round(sum(vals) / len(vals), 3) if vals else None
    return {
        "n": n,
        "wins": wins,
        "losses": losses,
        "ties": ties,
        "tie_half_win_rate": round(tie_half_wins / n, 4) if n else 0.0,
        "wilson_95_ci": [round(low, 4), round(high, 4)],
        "mean_confidence": round(sum(confidence_values) / len(confidence_values), 3) if confidence_values else None,
        "mean_familiarity": round(sum(familiarity_values) / len(familiarity_values), 3) if familiarity_values else None,
        "mean_dimension_delta": dim_means,
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = [
        "reviewer",
        "task",
        "item_id",
        "family",
        "private_comparison",
        "preference",
        "focused_full_side",
        "focused_full_result",
        "preferred_method",
        "confidence",
        "familiarity",
        "rationale",
        "concerns",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field) for field in fieldnames})


def write_markdown(path: Path, decoded: list[dict], summary: dict, errors: list[str]) -> None:
    lines = [
        "# Human Expert Blind Review Summary",
        "",
        "This summary decodes completed human review workbooks using the private answer key. "
        "It should be reported as a small expert sanity check, not as a large-scale human study.",
        "",
        "## Overall",
        "",
        "| N | focused_full wins | losses | ties | tie-half win rate | 95% Wilson CI | mean confidence | mean familiarity |",
        "| ---: | ---: | ---: | ---: | ---: | --- | ---: | ---: |",
    ]
    overall = summary["overall"]
    lines.append(
        f"| {overall['n']} | {overall['wins']} | {overall['losses']} | {overall['ties']} | "
        f"{overall['tie_half_win_rate']:.1%} | [{overall['wilson_95_ci'][0]:.1%}, {overall['wilson_95_ci'][1]:.1%}] | "
        f"{overall['mean_confidence']} | {overall['mean_familiarity']} |"
    )
    lines.extend(["", "## By Task", "", "| task | N | W/L/T | win rate | 95% CI | mean confidence |", "| --- | ---: | --- | ---: | --- | ---: |"])
    for task, item in sorted(summary["by_task"].items()):
        lines.append(
            f"| {task} | {item['n']} | {item['wins']}/{item['losses']}/{item['ties']} | "
            f"{item['tie_half_win_rate']:.1%} | [{item['wilson_95_ci'][0]:.1%}, {item['wilson_95_ci'][1]:.1%}] | "
            f"{item['mean_confidence']} |"
        )
    lines.extend(["", "## By Comparison", "", "| comparison | N | W/L/T | win rate | mean confidence |", "| --- | ---: | --- | ---: | ---: |"])
    for comp, item in sorted(summary["by_comparison"].items()):
        lines.append(
            f"| {comp} | {item['n']} | {item['wins']}/{item['losses']}/{item['ties']} | "
            f"{item['tie_half_win_rate']:.1%} | {item['mean_confidence']} |"
        )
    lines.extend(
        [
            "",
            "## Interpretation Notes",
            "",
            "- Current human evaluation has only one expert per completed domain, so it cannot establish inter-rater reliability.",
            "- Confidence is part of the result and should be reported; low-confidence preferences should not be overstated.",
            "- Missing domains should be reported as pending rather than silently ignored.",
            "- This expert check is best used together with the larger multi-LLM blind review, position-swap controls, ablations, and claim verification.",
            "",
        ]
    )
    if errors:
        lines.extend(["## Validation Warnings", ""])
        for error in errors:
            lines.append(f"- {error}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Summarize completed Chinese human review workbooks.")
    parser.add_argument("--answer-key", type=Path, default=Path("aaai27/human_evaluation/private_human_answer_key.jsonl"))
    parser.add_argument("--input", action="append", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=Path("aaai27/human_evaluation/results/human_review_summary_v1"))
    args = parser.parse_args()

    answer_key = {row["item_id"]: row for row in read_jsonl(args.answer_key)}
    workbook_rows = []
    for path in args.input:
        workbook_rows.extend(read_workbook(path))
    decoded, errors = decode_rows(workbook_rows, answer_key)

    summary = {
        "overall": summarize_group(decoded),
        "by_task": {},
        "by_comparison": {},
    }
    by_task = defaultdict(list)
    by_comparison = defaultdict(list)
    for row in decoded:
        by_task[row["task"]].append(row)
        by_comparison[row["private_comparison"]].append(row)
    summary["by_task"] = {key: summarize_group(value) for key, value in by_task.items()}
    summary["by_comparison"] = {key: summarize_group(value) for key, value in by_comparison.items()}

    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "human_review_summary.json").write_text(
        json.dumps({"summary": summary, "validation_errors": errors}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (args.output_dir / "decoded_human_review_results.jsonl").write_text(
        "\n".join(json.dumps(row, ensure_ascii=False) for row in decoded) + ("\n" if decoded else ""),
        encoding="utf-8",
    )
    write_csv(args.output_dir / "decoded_human_review_results.csv", decoded)
    write_markdown(args.output_dir / "HUMAN_EXPERT_BLIND_REVIEW_SUMMARY.md", decoded, summary, errors)

    print(f"Rows decoded: {len(decoded)}")
    print(f"Validation errors: {len(errors)}")
    print(f"Wrote {args.output_dir / 'HUMAN_EXPERT_BLIND_REVIEW_SUMMARY.md'}")
    print(
        "Overall:",
        f"W/L/T={summary['overall']['wins']}/{summary['overall']['losses']}/{summary['overall']['ties']}",
        f"win_rate={summary['overall']['tie_half_win_rate']:.3f}",
        f"mean_confidence={summary['overall']['mean_confidence']}",
    )


if __name__ == "__main__":
    main()
