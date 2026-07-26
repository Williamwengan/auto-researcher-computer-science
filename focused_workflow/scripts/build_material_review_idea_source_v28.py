#!/usr/bin/env python3
"""Parse the material expert review xlsx into a web/backend friendly JSON.

The demo originally used V10 final-plan templates. This script turns the human
expert review sheet into a first-class idea source so the material/physical page
can display the reviewed winner rather than the older V10 summary idea.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT / "material评审答题表_中文版.xlsx"
DEFAULT_OUT = ROOT / "competition_submission" / "material_review_ideas.json"


SCORE_DIMS = [
    "创新性",
    "研究吸引力",
    "可行性",
    "预期有效性",
    "综合质量",
    "基线依据充分性",
    "实验严谨性",
    "机制具体性",
    "实现就绪度",
]


def relpath(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT))
    except ValueError:
        return str(path)


def avg(values: list[Any]) -> float:
    nums = [float(v) for v in values if isinstance(v, (int, float))]
    return sum(nums) / len(nums) if nums else 0.0


def extract_title(text: str, fallback: str) -> str:
    patterns = [
        r"^Title:\s*\n(.+)$",
        r"^Idea\s+\d+\s+Title:\s*(.+)$",
        r"^(.+)$",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.MULTILINE)
        if m:
            title = m.group(1).strip()
            if title and len(title) < 180:
                return title
    return fallback


def extract_section(text: str, heading: str) -> str:
    pattern = rf"{re.escape(heading)}:\s*\n(.+?)(?=\n\n[A-Z][A-Za-z, /-]+:|\n\nEvidence paper IDs:|\n\nRisks, controls, or fallback:|\Z)"
    m = re.search(pattern, text, flags=re.DOTALL)
    return m.group(1).strip() if m else ""


def parse_workbook(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True)
    score_ws = wb["评分表"]
    content_ws = wb["候选内容"]

    score_headers = [c.value for c in score_ws[1]]
    score_rows: list[dict[str, Any]] = []
    for row in score_ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        score_rows.append(dict(zip(score_headers, row)))

    content_headers = [c.value for c in content_ws[1]]
    content_by_id: dict[str, dict[str, Any]] = {}
    for row in content_ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        d = dict(zip(content_headers, row))
        content_by_id[str(d.get("条目ID"))] = d

    reviewed: list[dict[str, Any]] = []
    for row in score_rows:
        item_id = str(row.get("条目ID"))
        content = content_by_id.get(item_id, {})
        a_score = avg([row.get(f"{dim}-A（1-5）") for dim in SCORE_DIMS])
        b_score = avg([row.get(f"{dim}-B（1-5）") for dim in SCORE_DIMS])
        pref = str(row.get("总体偏好（A/B/tie）") or "").strip()
        if pref == "A":
            winner = "A"
        elif pref == "B":
            winner = "B"
        else:
            winner = "A" if a_score >= b_score else "B"
        text = str(content.get(f"候选方案 {winner}（科研原文）") or "")
        title = extract_title(text, item_id)
        reviewed.append(
            {
                "item_id": item_id,
                "comparison_type": row.get("对比类型"),
                "winner": winner,
                "preference": pref or "score_based",
                "confidence": row.get("置信度（1-5）"),
                "score_a": round(a_score, 3),
                "score_b": round(b_score, 3),
                "winner_score": round(a_score if winner == "A" else b_score, 3),
                "title": title,
                "idea_text": text,
                "core_proposal": extract_section(text, "Core proposal"),
                "motivation": extract_section(text, "Motivation or baseline weakness"),
                "mechanism": extract_section(text, "Mechanism or approach"),
                "experiment_plan": extract_section(text, "Experiment and implementation plan"),
                "evidence_paper_ids": extract_section(text, "Evidence paper IDs"),
                "risks": extract_section(text, "Risks, controls, or fallback"),
                "review_reason": row.get("选择理由（必填，1-3句）"),
                "review_concern": row.get("疑虑或共同缺陷（选填）"),
            }
        )

    reviewed.sort(key=lambda x: (x.get("winner_score") or 0, x.get("confidence") or 0), reverse=True)
    best = reviewed[0] if reviewed else {}
    return {
        "version": "v28_material_human_review_idea_source",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": relpath(path),
        "reviewer_count_note": "Current sheet appears to contain one physical_property_expert reviewer; use as human-review signal, not high-confidence benchmark.",
        "num_scored_rows": len(score_rows),
        "num_candidate_rows": len(content_by_id),
        "best_idea": best,
        "reviewed_ideas": reviewed,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    xlsx = args.xlsx if args.xlsx.is_absolute() else ROOT / args.xlsx
    output = args.output if args.output.is_absolute() else ROOT / args.output
    if not xlsx.exists():
        raise FileNotFoundError(xlsx)
    data = parse_workbook(xlsx)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {output}")
    print(f"Best: {data.get('best_idea', {}).get('title')}")
    print(f"Rows: scored={data['num_scored_rows']} candidates={data['num_candidate_rows']}")


if __name__ == "__main__":
    main()
